"""Excel file handling utilities"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from utils.constants import (
    EXCEL_USERS_PATH,
    EXCEL_TEST_RESULTS_PATH,
    BASE_DATA_PATH
)


class ExcelHandler:
    """Handles all Excel operations"""

    @staticmethod
    def ensure_data_folder():
        """Ensure data folder exists"""
        os.makedirs(BASE_DATA_PATH, exist_ok=True)

    @staticmethod
    def init_users_excel():
        """Initialize users credentials Excel file"""
        ExcelHandler.ensure_data_folder()
        
        if not os.path.exists(EXCEL_USERS_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            
            # Headers
            headers = ["Username", "Password", "Registration Date", "Last Login"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(EXCEL_USERS_PATH)

    @staticmethod
    def init_test_results_excel():
        """Initialize test results Excel file"""
        ExcelHandler.ensure_data_folder()
        
        if not os.path.exists(EXCEL_TEST_RESULTS_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Headers
            headers = ["Username", "Grade", "Subject", "Topic", "Paper Number", 
                      "Questions Answered", "Correct Answers", "Score %", "Test Date"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(EXCEL_TEST_RESULTS_PATH)

    @staticmethod
    def register_user(username: str, password: str) -> bool:
        """Register a new user"""
        try:
            ExcelHandler.init_users_excel()
            wb = load_workbook(EXCEL_USERS_PATH)
            ws = wb.active
            
            # Check if user already exists
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == username:
                    return False  # User already exists
            
            # Add new user
            ws.append([username, password, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "-"])
            wb.save(EXCEL_USERS_PATH)
            return True
        except Exception as e:
            print(f"Error registering user: {e}")
            return False

    @staticmethod
    def validate_user(username: str, password: str) -> bool:
        """Validate user credentials"""
        try:
            if not os.path.exists(EXCEL_USERS_PATH):
                return False
            
            wb = load_workbook(EXCEL_USERS_PATH)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == username and row[1] == password:
                    return True
            
            return False
        except Exception as e:
            print(f"Error validating user: {e}")
            return False

    @staticmethod
    def user_exists(username: str) -> bool:
        """Check if user already exists"""
        try:
            if not os.path.exists(EXCEL_USERS_PATH):
                return False
            
            wb = load_workbook(EXCEL_USERS_PATH)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == username:
                    return True
            
            return False
        except Exception as e:
            print(f"Error checking user existence: {e}")
            return False

    @staticmethod
    def save_test_result(username: str, grade: str, subject: str, topic: str, 
                        paper_num: int, total_questions: int, correct_answers: int) -> bool:
        """Save test result to Excel"""
        try:
            ExcelHandler.init_test_results_excel()
            wb = load_workbook(EXCEL_TEST_RESULTS_PATH)
            ws = wb.active
            
            score_percentage = (correct_answers / total_questions) * 100
            
            ws.append([
                username,
                grade,
                subject,
                topic,
                paper_num,
                total_questions,
                correct_answers,
                f"{score_percentage:.2f}%",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
            
            wb.save(EXCEL_TEST_RESULTS_PATH)
            return True
        except Exception as e:
            print(f"Error saving test result: {e}")
            return False

    @staticmethod
    def create_questions_excel(grade: str, subject: str, topic: str) -> str:
        """Create Excel file for questions"""
        filename = f"questions_{grade}_{subject}_{topic}.xlsx"
        filepath = os.path.join(BASE_DATA_PATH, filename)
        
        ExcelHandler.ensure_data_folder()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Questions"
        
        # Headers
        headers = ["Question #", "Question Text", "Option A", "Option B", "Option C", "Option D", "Correct Answer"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        wb.save(filepath)
        return filepath

    @staticmethod
    def add_questions_to_excel(filepath: str, questions: list) -> bool:
        """Add questions to Excel file"""
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            for i, q in enumerate(questions, 1):
                ws.append([
                    i,
                    q.get("question", ""),
                    q.get("options", [None, None, None, None])[0],
                    q.get("options", [None, None, None, None])[1],
                    q.get("options", [None, None, None, None])[2],
                    q.get("options", [None, None, None, None])[3],
                    q.get("answer", "")
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Error adding questions to Excel: {e}")
            return False

    @staticmethod
    def get_student_results(username: str) -> list:
        """Get all test results for a specific student"""
        try:
            if not os.path.exists(EXCEL_TEST_RESULTS_PATH):
                return []
            
            wb = load_workbook(EXCEL_TEST_RESULTS_PATH)
            ws = wb.active
            
            results = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == username:
                    results.append(row)
            
            return results
        except Exception as e:
            print(f"Error retrieving student results: {e}")
            return []